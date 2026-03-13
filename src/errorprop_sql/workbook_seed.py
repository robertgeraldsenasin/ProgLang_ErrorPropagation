from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any
import re

import openpyxl
import yaml

from .config import ERROR_SEVERITY, FEEDBACK_DESCRIPTIONS
from .task_loader import get_task_by_id
from .utils import safe_model_slug
from .workbook_sync import (
    RUN_PLAN_HEADERS,
    TASK_HEADERS,
    _clear_row_values,
    _copy_row_style,
    _first_blank_row,
    _set_dashboard_formulas,
)


def _mark_for_recalc(wb) -> None:
    calc = getattr(wb, "calculation", None)
    if calc is not None:
        calc.fullCalcOnLoad = True
        calc.forceFullCalc = True


@dataclass
class PlannedRun:
    run_id: str
    batch_id: str
    task_id: str
    db_name: str
    model_id: str
    model_snapshot: str
    protocol_id: str
    temperature: float
    reasoning_mode: str
    T_max: int
    replicate: int
    operator: str
    planned_date: str
    record_file: str
    artifact_folder: str
    commit_hash: str
    status: str
    notes: str


def _read_yaml(path: Path) -> dict[str, Any]:
    return yaml.safe_load(path.read_text(encoding="utf-8")) or {}


def _replace_table_rows(ws, start_row: int, values: list[list[Any]], max_cols: int) -> None:
    for row in range(start_row, max(ws.max_row, start_row) + 1):
        _clear_row_values(ws, row, max_cols)
    for offset, row_values in enumerate(values, start=start_row):
        if offset > ws.max_row:
            _copy_row_style(ws, start_row, offset, max_cols)
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(offset, col_idx).value = value


def _version_label(prompt_text: str) -> str:
    return "v2" if "{supporting_context}" in prompt_text else "v1"


def _extract_required_variables(prompt_text: str) -> str:
    vars_found = sorted(set(re.findall(r"{([^}]+)}", prompt_text)))
    return ", ".join(f"{{{v}}}" for v in vars_found)


def _seed_lookups(wb, model_suite: dict[str, Any]) -> None:
    ws = wb["Lookups"]
    state_rows = [[state, severity] for state, severity in ERROR_SEVERITY.items()]
    protocol_rows = [[pid, desc] for pid, desc in FEEDBACK_DESCRIPTIONS.items()]
    model_rows = [[row["model_id"], row["provider"]] for row in model_suite.get("models", [])]

    for idx, row_values in enumerate(state_rows, start=2):
        ws.cell(idx, 1).value = row_values[0]
        ws.cell(idx, 2).value = row_values[1]

    for row in range(2, max(ws.max_row, 20) + 1):
        if row - 2 >= len(state_rows):
            ws.cell(row, 1).value = None
            ws.cell(row, 2).value = None
        if row - 2 >= len(protocol_rows):
            ws.cell(row, 4).value = None
            ws.cell(row, 5).value = None
        if row - 2 >= len(model_rows):
            ws.cell(row, 7).value = None
            ws.cell(row, 8).value = None

    for idx, row_values in enumerate(protocol_rows, start=2):
        ws.cell(idx, 4).value = row_values[0]
        ws.cell(idx, 5).value = row_values[1]

    for idx, row_values in enumerate(model_rows, start=2):
        ws.cell(idx, 7).value = row_values[0]
        ws.cell(idx, 8).value = row_values[1]


def _seed_prompt_library(repo_root: Path, wb) -> None:
    ws = wb["Prompt Library"]
    prompt_map = {
        "T1_SQL": ("Turn 1 / initial generation", repo_root / "prompts" / "turn1_sql.txt"),
        "REV_F0": ("Turn k / Minimal feedback", repo_root / "prompts" / "revise_F0.txt"),
        "REV_F1": ("Turn k / Engine error feedback", repo_root / "prompts" / "revise_F1.txt"),
        "REV_F2": ("Turn k / Output-difference feedback", repo_root / "prompts" / "revise_F2.txt"),
        "REV_F3": ("Turn k / Self-diagnosis + correction", repo_root / "prompts" / "revise_F3.txt"),
    }
    rows: list[list[Any]] = []
    for template_id, (used_when, path) in prompt_map.items():
        text = path.read_text(encoding="utf-8").strip()
        rows.append(
            [
                template_id,
                used_when,
                _version_label(text),
                text,
                _extract_required_variables(text),
                f"Sourced from prompts/{path.name}",
            ]
        )
    _replace_table_rows(ws, 5, rows, 6)


def _task_row_from_spider2(spider2_root: Path, task_id: str, group_name: str) -> list[Any]:
    task = get_task_by_id(spider2_root, task_id)
    oracle_exec = spider2_root / "spider2-lite" / "evaluation_suite" / "gold" / "exec_result"
    oracle_sql = spider2_root / "spider2-lite" / "evaluation_suite" / "gold" / "sql"
    oracle_path = ""
    matches = sorted(oracle_exec.glob(f"{task_id}.*")) if oracle_exec.exists() else []
    if matches:
        oracle_path = str(matches[0].relative_to(spider2_root))
    else:
        sql_path = oracle_sql / f"{task_id}.sql"
        if sql_path.exists():
            oracle_path = str(sql_path.relative_to(spider2_root))

    return [
        task.instance_id,
        "spider2-lite-sqlite",
        task.db,
        f"Spider2/spider2-lite/resource/databases/spider2-localdb/{task.db}.sqlite",
        task.question,
        "",
        oracle_path,
        "",
        "Yes" if group_name == "pilot" else "No",
        "Yes" if group_name == "main" else "No",
        group_name.upper(),
        str(task.external_knowledge or ""),
    ]


def _task_row_fallback(task_id: str, group_name: str, metadata: dict[str, Any] | None = None) -> list[Any]:
    metadata = metadata or {}
    db_name = metadata.get("db_name", "")
    question = metadata.get("question", "")
    external = metadata.get("external_knowledge", "")
    db_path = f"Spider2/spider2-lite/resource/databases/spider2-localdb/{db_name}.sqlite" if db_name else ""
    notes = str(external or "Populate oracle path by rerunning this script with --spider2-root")
    return [
        task_id,
        "spider2-lite-sqlite",
        db_name,
        db_path,
        question,
        "",
        "",
        "",
        "Yes" if group_name == "pilot" else "No",
        "Yes" if group_name == "main" else "No",
        group_name.upper(),
        notes,
    ]


def _seed_tasks(wb, spider2_root: Path | None, task_pack: dict[str, Any], include_holdout: bool) -> None:
    ws = wb["Tasks"]
    rows: list[list[Any]] = []
    groups = task_pack.get("groups", {})
    task_metadata = task_pack.get("task_metadata", {})
    ordered_groups = ["pilot", "main"] + (["holdout"] if include_holdout else []) + ["reserve"]
    for group_name in ordered_groups:
        for task_id in groups.get(group_name, []):
            if spider2_root:
                try:
                    rows.append(_task_row_from_spider2(spider2_root, task_id, group_name))
                except Exception:
                    rows.append(_task_row_fallback(task_id, group_name, task_metadata.get(task_id)))
            else:
                rows.append(_task_row_fallback(task_id, group_name, task_metadata.get(task_id)))
    _replace_table_rows(ws, 5, rows, len(TASK_HEADERS))


def _build_planned_runs(
    wb,
    spider2_root: Path | None,
    model_suite: dict[str, Any],
    task_pack: dict[str, Any],
    protocol_cfg: dict[str, Any],
    include_holdout: bool,
) -> list[PlannedRun]:
    models = model_suite.get("models", [])
    pilot_ids = task_pack.get("groups", {}).get("pilot", [])
    main_ids = task_pack.get("groups", {}).get("main", [])
    holdout_ids = task_pack.get("groups", {}).get("holdout", []) if include_holdout else []
    task_metadata = task_pack.get("task_metadata", {})
    runs: list[PlannedRun] = []

    def resolve_db(task_id: str) -> str:
        if spider2_root:
            try:
                return get_task_by_id(spider2_root, task_id).db
            except Exception:
                pass
        return str(task_metadata.get(task_id, {}).get("db_name", ""))

    def add_runs(task_ids: list[str], batch_id: str, protocol_id: str) -> None:
        for task_id in task_ids:
            db_name = resolve_db(task_id)
            for model in models:
                model_id = model["model_id"]
                run_id = f"{task_id.upper()}-{safe_model_slug(model_id).upper()}-{protocol_id}-R1"
                runs.append(
                    PlannedRun(
                        run_id=run_id,
                        batch_id=batch_id,
                        task_id=task_id,
                        db_name=db_name,
                        model_id=model_id,
                        model_snapshot=model.get("visible_label", model_id),
                        protocol_id=protocol_id,
                        temperature=0.0,
                        reasoning_mode="manual-browser",
                        T_max=int(protocol_cfg.get("turn_cap", 4)),
                        replicate=1,
                        operator="",
                        planned_date="",
                        record_file="",
                        artifact_folder=f"runs/{run_id}",
                        commit_hash="",
                        status="Planned",
                        notes=f"Seeded from {batch_id.lower()} pack",
                    )
                )

    add_runs(pilot_ids, "PILOT-F1", str(protocol_cfg.get("pilot_protocol", "F1")))
    add_runs(main_ids, "MAIN-F2", str(protocol_cfg.get("main_protocol", "F2")))
    if holdout_ids:
        add_runs(holdout_ids, "HOLDOUT-F2", str(protocol_cfg.get("holdout_protocol", "F2")))
    return runs


def _seed_run_plan(wb, runs: list[PlannedRun]) -> None:
    ws = wb["Run Plan"]
    rows = [[getattr(run, header) for header in RUN_PLAN_HEADERS] for run in runs]
    _replace_table_rows(ws, 5, rows, len(RUN_PLAN_HEADERS))


def seed_workbook_reference_data(
    workbook_path: Path,
    *,
    repo_root: Path,
    model_suite_path: Path,
    task_pack_path: Path,
    study_protocol_path: Path,
    spider2_root: Path | None = None,
    include_holdout: bool = False,
) -> None:
    model_suite = _read_yaml(model_suite_path)
    task_pack = _read_yaml(task_pack_path)
    protocol_cfg = _read_yaml(study_protocol_path)

    wb = openpyxl.load_workbook(workbook_path)
    _seed_lookups(wb, model_suite)
    _seed_prompt_library(repo_root, wb)
    _seed_tasks(wb, spider2_root, task_pack, include_holdout)
    planned_runs = _build_planned_runs(wb, spider2_root, model_suite, task_pack, protocol_cfg, include_holdout)
    _seed_run_plan(wb, planned_runs)
    _set_dashboard_formulas(wb["Dashboard"])

    _mark_for_recalc(wb)
    wb.save(workbook_path)
