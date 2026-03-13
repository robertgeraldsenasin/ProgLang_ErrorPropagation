from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path
import subprocess
from typing import Any

import openpyxl

from .utils import safe_model_slug


def _mark_for_recalc(wb) -> None:
    calc = getattr(wb, "calculation", None)
    if calc is not None:
        calc.fullCalcOnLoad = True
        calc.forceFullCalc = True


@dataclass
class WorkbookRunRequest:
    row_index: int
    run_id: str
    batch_id: str
    task_id: str
    db_name: str
    model_id: str
    model_snapshot: str
    protocol_id: str
    temperature: float
    reasoning_mode: str
    t_max: int
    replicate: int
    operator: str
    planned_date: str
    record_file: str
    artifact_folder: str
    commit_hash: str
    status: str
    notes: str


def _header_map(ws) -> dict[str, int]:
    return {
        str(ws.cell(4, col).value): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(4, col).value
    }


def build_run_id(task_id: str, model_id: str, protocol_id: str, replicate: int) -> str:
    return f"{task_id.upper()}-{safe_model_slug(model_id).upper()}-{protocol_id}-R{replicate}"


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _float(value: Any, default: float) -> float:
    if value in (None, ""):
        return default
    return float(value)


def _int(value: Any, default: int) -> int:
    if value in (None, ""):
        return default
    return int(value)


def current_git_commit(repo_root: Path) -> str:
    try:
        result = subprocess.run(
            ["git", "rev-parse", "--short", "HEAD"],
            cwd=repo_root,
            check=True,
            capture_output=True,
            text=True,
        )
        return result.stdout.strip()
    except Exception:
        return ""


def load_run_request(workbook_path: Path, *, row: int | None = None, run_id: str | None = None) -> WorkbookRunRequest:
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb["Run Plan"]
    headers = _header_map(ws)

    if row is None and not run_id:
        raise ValueError("Provide either row or run_id.")
    if row is None:
        found_row = None
        for candidate in range(5, ws.max_row + 1):
            if _text(ws.cell(candidate, headers["run_id"]).value) == str(run_id):
                found_row = candidate
                break
        if found_row is None:
            raise KeyError(f"Run ID not found in workbook: {run_id}")
        row = found_row

    def cell(name: str):
        return ws.cell(row, headers[name]).value

    task_id = _text(cell("task_id"))
    model_id = _text(cell("model_id"))
    protocol_id = _text(cell("protocol_id"))
    if not task_id or not model_id or not protocol_id:
        raise ValueError("Run Plan row must include task_id, model_id, and protocol_id before execution.")

    replicate = _int(cell("replicate"), 1)
    resolved_run_id = _text(cell("run_id")) or build_run_id(task_id, model_id, protocol_id, replicate)

    return WorkbookRunRequest(
        row_index=row,
        run_id=resolved_run_id,
        batch_id=_text(cell("batch_id")) or "MAIN",
        task_id=task_id,
        db_name=_text(cell("db_name")),
        model_id=model_id,
        model_snapshot=_text(cell("model_snapshot")) or model_id,
        protocol_id=protocol_id,
        temperature=_float(cell("temperature"), 0.0),
        reasoning_mode=_text(cell("reasoning_mode")) or "manual-browser",
        t_max=_int(cell("T_max"), 4),
        replicate=replicate,
        operator=_text(cell("operator")),
        planned_date=_text(cell("planned_date")) or date.today().isoformat(),
        record_file=_text(cell("record_file")),
        artifact_folder=_text(cell("artifact_folder")) or f"runs/{resolved_run_id}",
        commit_hash=_text(cell("commit_hash")),
        status=_text(cell("status")) or "Planned",
        notes=_text(cell("notes")),
    )


def update_run_plan_row(workbook_path: Path, row_index: int, updates: dict[str, Any]) -> None:
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb["Run Plan"]
    headers = _header_map(ws)
    for key, value in updates.items():
        if key not in headers or value is None:
            continue
        ws.cell(row_index, headers[key]).value = value
    _mark_for_recalc(wb)
    wb.save(workbook_path)
