from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Iterable

import openpyxl

from .task_loader import load_tasks_from_root
from .utils import read_jsonl

TURN_LOG_HEADERS = [
    "batch_id",
    "run_id",
    "turn_no",
    "task_id",
    "db_name",
    "model_id",
    "protocol_id",
    "prompt_file",
    "response_file",
    "extracted_sql_file",
    "prompt_hash",
    "exec_ms",
    "rows_pred",
    "rows_gold",
    "symdiff_rows",
    "sqlite_error",
    "exec_state",
    "severity",
    "helper_key",
    "helper_prev_key",
    "prev_state",
    "prev_severity",
    "state_change",
    "improved",
    "regressed",
    "pass_flag",
    "feedback_file",
    "screenshot_file",
    "recording_timestamp",
    "notes",
    "review_flag",
]

RUN_PLAN_HEADERS = [
    "run_id",
    "batch_id",
    "task_id",
    "db_name",
    "model_id",
    "model_snapshot",
    "protocol_id",
    "temperature",
    "reasoning_mode",
    "T_max",
    "replicate",
    "operator",
    "planned_date",
    "record_file",
    "artifact_folder",
    "commit_hash",
    "status",
    "notes",
]

TASK_HEADERS = [
    "task_id",
    "dataset_split",
    "db_name",
    "db_path",
    "question",
    "schema_hash / version",
    "oracle_result_file",
    "difficulty_band",
    "pilot?",
    "main?",
    "recording_segment",
    "notes",
]

STABILITY_HEADERS = [
    "run_id",
    "pass_turn",
    "check_type",
    "repeat_no",
    "outcome_state",
    "stable_pass?",
    "evidence_file",
    "notes",
]


def _mark_for_recalc(wb) -> None:
    calc = getattr(wb, "calculation", None)
    if calc is not None:
        calc.fullCalcOnLoad = True
        calc.forceFullCalc = True


def _first_blank_row(ws, key_col: int, start_row: int = 5) -> int:
    row = start_row
    while ws.cell(row, key_col).value not in (None, ""):
        row += 1
    return row


def _last_populated_row(ws, key_col: int, start_row: int = 5) -> int:
    last = start_row - 1
    for row in range(start_row, ws.max_row + 1):
        value = ws.cell(row, key_col).value
        if value not in (None, ""):
            last = row
    return last


def _copy_row_style(ws, source_row: int, target_row: int, max_col: int) -> None:
    for c in range(1, max_col + 1):
        src = ws.cell(source_row, c)
        dst = ws.cell(target_row, c)
        if src.has_style:
            dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def _clear_row_values(ws, row: int, max_col: int, start_col: int = 1) -> None:
    for c in range(start_col, max_col + 1):
        ws.cell(row, c).value = None


def _clear_rows_by_prefix(ws, key_col: int, prefixes: Iterable[str], max_col: int, start_row: int = 5) -> None:
    prefixes = tuple(prefixes)
    for row in range(start_row, ws.max_row + 1):
        value = ws.cell(row, key_col).value
        if isinstance(value, str) and value.startswith(prefixes):
            _clear_row_values(ws, row, max_col)


def _formulaify_turn_log_row(ws, row: int) -> None:
    ws[f"R{row}"] = f'=IF(Q{row}="","",IFERROR(VLOOKUP(Q{row},Lookups!$A$2:$B$8,2,FALSE),""))'
    ws[f"S{row}"] = f'=IF(B{row}="","",B{row}&"|"&TEXT(C{row},"00"))'
    ws[f"T{row}"] = f'=IF(B{row}="","",B{row}&"|"&TEXT(C{row}-1,"00"))'
    ws[f"U{row}"] = f'=IFERROR(INDEX($Q$5:$Q$5000,MATCH(T{row},$S$5:$S$5000,0)),"")'
    ws[f"V{row}"] = f'=IFERROR(INDEX($R$5:$R$5000,MATCH(T{row},$S$5:$S$5000,0)),"")'
    ws[f"W{row}"] = f'=IF(U{row}="","",IF(Q{row}<>U{row},1,0))'
    ws[f"X{row}"] = f'=IF(V{row}="","",IF(R{row}<V{row},1,0))'
    ws[f"Y{row}"] = f'=IF(V{row}="","",IF(R{row}>V{row},1,0))'
    ws[f"Z{row}"] = f'=IF(Q{row}="","",IF(Q{row}="Pass",1,0))'


def _has_informative_value(value) -> bool:
    if value is None:
        return False
    if isinstance(value, str) and value.strip() == "":
        return False
    return True


def _upsert_rows(
    ws,
    rows: list[dict],
    headers: list[str],
    dedupe_keys: list[str],
    key_col: int,
    style_source_row: int,
) -> None:
    row_index: dict[tuple, int] = {}
    header_to_col = {header: idx + 1 for idx, header in enumerate(headers)}

    for r in range(5, ws.max_row + 1):
        first = ws.cell(r, key_col).value
        if first in (None, ""):
            continue
        dedupe = tuple(ws.cell(r, header_to_col[k]).value for k in dedupe_keys)
        row_index[dedupe] = r

    for row in rows:
        dedupe = tuple(row.get(k) for k in dedupe_keys)
        target = row_index.get(dedupe)
        if target is None:
            target = _first_blank_row(ws, key_col)
            _copy_row_style(ws, style_source_row, target, len(headers))
            row_index[dedupe] = target

        for i, header in enumerate(headers, start=1):
            value = row.get(header)
            if _has_informative_value(value):
                ws.cell(target, i).value = value


def _set_run_summary_row_formulas(ws, row: int) -> None:
    formulas_by_col = {
        "A": "=IF('Run Plan'!A{r}=\"\",\"\",'Run Plan'!A{r})",
        "B": "=IF(A{r}=\"\",\"\",'Run Plan'!C{r})",
        "C": "=IF(A{r}=\"\",\"\",'Run Plan'!E{r})",
        "D": "=IF(A{r}=\"\",\"\",'Run Plan'!G{r})",
        "E": "=IF(A{r}=\"\",\"\",'Run Plan'!K{r})",
        "F": "=IF(A{r}=\"\",\"\",'Run Plan'!J{r})",
        "G": "=IF(A{r}=\"\",\"\",COUNTIF('Turn Log'!$B$5:$B$5000,A{r}))",
        "H": "=IF(A{r}=\"\",\"\",IFERROR(MINIFS('Turn Log'!$C$5:$C$5000,'Turn Log'!$B$5:$B$5000,A{r},'Turn Log'!$Q$5:$Q$5000,\"Pass\"),\"\"))",
        "I": "=IF(A{r}=\"\",\"\",IFERROR(MAXIFS('Turn Log'!$C$5:$C$5000,'Turn Log'!$B$5:$B$5000,A{r}),\"\"))",
        "J": "=IF(A{r}=\"\",\"\",IFERROR(INDEX('Turn Log'!$Q$5:$Q$5000,MATCH(A{r}&\"|\"&TEXT(I{r},\"00\"),'Turn Log'!$S$5:$S$5000,0)),\"\"))",
        "K": "=IF(A{r}=\"\",\"\",IFERROR(INDEX('Turn Log'!$R$5:$R$5000,MATCH(A{r}&\"|\"&TEXT(I{r},\"00\"),'Turn Log'!$S$5:$S$5000,0)),\"\"))",
        "L": "=IF(A{r}=\"\",\"\",COUNTIFS('Turn Log'!$B$5:$B$5000,A{r},'Turn Log'!$X$5:$X$5000,1))",
        "M": "=IF(A{r}=\"\",\"\",COUNTIFS('Turn Log'!$B$5:$B$5000,A{r},'Turn Log'!$Y$5:$Y$5000,1))",
        "N": "=IF(A{r}=\"\",\"\",COUNTIFS('Turn Log'!$B$5:$B$5000,A{r},'Turn Log'!$W$5:$W$5000,1))",
        "O": "=IF(A{r}=\"\",\"\",IF(H{r}=\"\",\"No\",\"Yes\"))",
        "P": "=IF(A{r}=\"\",\"\",COUNTIFS('Stability Checks'!$A$5:$A$5000,A{r},'Stability Checks'!$C$5:$C$5000,\"Re-execution\",'Stability Checks'!$E$5:$E$5000,\"Pass\"))",
        "Q": "=IF(A{r}=\"\",\"\",COUNTIFS('Stability Checks'!$A$5:$A$5000,A{r},'Stability Checks'!$C$5:$C$5000,\"Re-prompt\",'Stability Checks'!$E$5:$E$5000,\"Pass\"))",
        "R": "=IF(A{r}=\"\",\"\",'Run Plan'!Q{r})",
    }
    for col_letter, template in formulas_by_col.items():
        ws[f"{col_letter}{row}"] = template.format(r=row)


def _ensure_run_summary_rows(ws, run_plan_ws) -> None:
    last_run_plan_row = _last_populated_row(run_plan_ws, 1)
    template_row = 5
    target_last_row = max(last_run_plan_row, template_row)

    for row in range(template_row, target_last_row + 1):
        if row > ws.max_row:
            _copy_row_style(ws, template_row, row, ws.max_column)
        _set_run_summary_row_formulas(ws, row)

    for row in range(target_last_row + 1, ws.max_row + 1):
        _clear_row_values(ws, row, 18)


def _set_dashboard_formulas(ws) -> None:
    ws["B4"] = '=COUNTIF(\'Run Plan\'!$A$5:$A$5000,"?*")'
    ws["D4"] = '=COUNTIF(\'Run Summary\'!$O$5:$O$5000,"Yes")'
    ws["F4"] = '=IFERROR(AVERAGEIF(\'Run Summary\'!$H$5:$H$5000,">0"),"")'
    ws["B8"] = '=IFERROR(AVERAGEIF(\'Run Summary\'!$A$5:$A$5000,"?*",\'Run Summary\'!$M$5:$M$5000),"")'
    ws["D8"] = '=COUNTIF(\'Run Summary\'!$A$5:$A$5000,"?*")'
    ws["F8"] = '=COUNTIF(\'Stability Checks\'!$A$5:$A$5000,"?*")'


def _oracle_info_for_task(spider2_root: Path, instance_id: str) -> tuple[str, str]:
    gold_exec_dir = spider2_root / "spider2-lite" / "evaluation_suite" / "gold" / "exec_result"
    gold_sql_dir = spider2_root / "spider2-lite" / "evaluation_suite" / "gold" / "sql"
    matches = sorted(gold_exec_dir.glob(f"{instance_id}.*")) if gold_exec_dir.exists() else []
    if matches:
        return "exec_result", str(matches[0].relative_to(spider2_root))
    sql_path = gold_sql_dir / f"{instance_id}.sql"
    if sql_path.exists():
        return "gold_sql", str(sql_path.relative_to(spider2_root))
    return "", ""


def populate_tasks_sheet_from_spider2(workbook_path: Path, spider2_root: Path, *, local_only: bool = True) -> None:
    wb = openpyxl.load_workbook(workbook_path)
    tasks_ws = wb["Tasks"]

    for row in range(5, max(tasks_ws.max_row, 5) + 1):
        _clear_row_values(tasks_ws, row, len(TASK_HEADERS))

    tasks = load_tasks_from_root(spider2_root)
    if local_only:
        tasks = [task for task in tasks if task.instance_id.lower().startswith("local")]
    tasks = sorted(tasks, key=lambda task: task.instance_id)

    for offset, task in enumerate(tasks, start=5):
        if offset > tasks_ws.max_row:
            _copy_row_style(tasks_ws, 5, offset, len(TASK_HEADERS))
        oracle_type, oracle_path = _oracle_info_for_task(spider2_root, task.instance_id)
        row = {
            "task_id": task.instance_id,
            "dataset_split": "spider2-lite-local",
            "db_name": task.db,
            "db_path": f"spider2-lite/resource/databases/spider2-localdb/{task.db}.sqlite",
            "question": task.question,
            "schema_hash / version": "",
            "oracle_result_file": oracle_path,
            "difficulty_band": "",
            "pilot?": "",
            "main?": "",
            "recording_segment": "",
            "notes": task.external_knowledge if isinstance(task.external_knowledge, str) else oracle_type,
        }
        for col, header in enumerate(TASK_HEADERS, start=1):
            tasks_ws.cell(offset, col).value = row.get(header, "")

    _mark_for_recalc(wb)
    wb.save(workbook_path)


def repair_workbook_layout(
    workbook_path: Path,
    *,
    drop_template_examples: bool = True,
    clear_task_rows: bool = False,
    clear_working_sheets: bool = False,
    spider2_root: Path | None = None,
    populate_tasks: bool = False,
) -> None:
    wb = openpyxl.load_workbook(workbook_path)
    run_plan_ws = wb["Run Plan"]
    turn_log_ws = wb["Turn Log"]
    stability_ws = wb["Stability Checks"]
    run_summary_ws = wb["Run Summary"]
    dashboard_ws = wb["Dashboard"]
    tasks_ws = wb["Tasks"]

    if drop_template_examples:
        _clear_rows_by_prefix(run_plan_ws, 1, ["PILOT-"], len(RUN_PLAN_HEADERS))
        _clear_rows_by_prefix(turn_log_ws, 2, ["PILOT-"], len(TURN_LOG_HEADERS))
        _clear_rows_by_prefix(stability_ws, 1, ["PILOT-"], len(STABILITY_HEADERS))

    if clear_task_rows and not populate_tasks:
        for row in range(5, max(tasks_ws.max_row, 5) + 1):
            _clear_row_values(tasks_ws, row, len(TASK_HEADERS))

    if clear_working_sheets:
        for row in range(5, max(turn_log_ws.max_row, 5) + 1):
            _clear_row_values(turn_log_ws, row, len(TURN_LOG_HEADERS))
        for row in range(5, max(stability_ws.max_row, 5) + 1):
            _clear_row_values(stability_ws, row, len(STABILITY_HEADERS))

    _ensure_run_summary_rows(run_summary_ws, run_plan_ws)
    _set_dashboard_formulas(dashboard_ws)

    _mark_for_recalc(wb)
    wb.save(workbook_path)

    if populate_tasks:
        if spider2_root is None:
            raise ValueError("populate_tasks=True requires spider2_root to be provided.")
        populate_tasks_sheet_from_spider2(workbook_path, spider2_root, local_only=True)
        wb = openpyxl.load_workbook(workbook_path)
        _set_dashboard_formulas(wb["Dashboard"])
        _mark_for_recalc(wb)
        wb.save(workbook_path)


def sync_output_to_workbook(workbook_path: Path, out_dir: Path) -> None:
    wb = openpyxl.load_workbook(workbook_path)
    run_plan_ws = wb["Run Plan"]
    turn_log_ws = wb["Turn Log"]
    stability_ws = wb["Stability Checks"]
    run_summary_ws = wb["Run Summary"]
    dashboard_ws = wb["Dashboard"]

    logs_dir = out_dir / "logs"
    run_plan_rows = read_jsonl(logs_dir / "run_plan.jsonl")
    turn_log_rows = read_jsonl(logs_dir / "turn_log.jsonl")
    stability_rows = read_jsonl(logs_dir / "stability_checks.jsonl")

    _upsert_rows(run_plan_ws, run_plan_rows, RUN_PLAN_HEADERS, ["run_id"], 1, 5)
    _upsert_rows(turn_log_ws, turn_log_rows, TURN_LOG_HEADERS, ["run_id", "turn_no"], 2, 7)
    _upsert_rows(stability_ws, stability_rows, STABILITY_HEADERS, ["run_id", "check_type", "repeat_no"], 1, 5)

    for r in range(5, turn_log_ws.max_row + 1):
        if turn_log_ws[f"B{r}"].value not in (None, ""):
            _formulaify_turn_log_row(turn_log_ws, r)

    _ensure_run_summary_rows(run_summary_ws, run_plan_ws)
    _set_dashboard_formulas(dashboard_ws)

    _mark_for_recalc(wb)
    wb.save(workbook_path)
