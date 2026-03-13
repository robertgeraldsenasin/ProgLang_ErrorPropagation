from __future__ import annotations

from pathlib import Path
import argparse
import openpyxl


RUN_PLAN_COLS = 18
TURN_LOG_COLS = 31
STABILITY_COLS = 8
TASK_COLS = 12
RUN_SUMMARY_COLS = 18


def mark_for_recalc(wb) -> None:
    calc = getattr(wb, "calculation", None)
    if calc is not None:
        calc.fullCalcOnLoad = True
        calc.forceFullCalc = True


def clear_rows(ws, start_row: int, max_col: int) -> None:
    for row in range(start_row, ws.max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row, col).value = None


def set_run_summary_row_formulas(ws, row: int) -> None:
    formulas_by_col = {
        "A": "=IF('Run Plan'!A{r}=\"\",\"\",'Run Plan'!A{r})",
        "B": "=IF(A{r}=\"\",\"\",'Run Plan'!C{r})",
        "C": "=IF(A{r}=\"\",\"\",'Run Plan'!E{r})",
        "D": "=IF(A{r}=\"\",\"\",'Run Plan'!G{r})",
        "E": "=IF(A{r}=\"\",\"\",'Run Plan'!K{r})",
        "F": "=IF(A{r}=\"\",\"\",'Run Plan'!J{r})",
        "G": "=IF(A{r}=\"\",\"\",COUNTIF('Turn Log'!$B$5:$B$5000,A{r}))",
        "H": "=IF(A{r}=\"\",\"\",IFERROR(MINIFS('Turn Log'!$C$5:$C$5000,'Turn Log'!$B$5:$B$5000,A{r},'Turn Log'!$Q$5:$Q$5000,\"Pass\"),\"\"))",
        "I": "=IF(A{r}=\"\",\"\",IFERROR(MAXIFS('Turn Log'!$C$5:$C$5000,'Turn Log'!$B$5:$B$5000,A{r}),\"\"))",
        "J": "=IF(A{r}=\"\",\"\",IFERROR(INDEX('Turn Log'!$Q$5:$Q$5000,MATCH(A{r}&\"|\"&TEXT(I{r},\"00\"),'Turn Log'!$S$5:$S$5000,0)),\"\"))",
        "K": "=IF(A{r}=\"\",\"\",IFERROR(INDEX('Turn Log'!$R$5:$R$5000,MATCH(A{r}&\"|\"&TEXT(I{r},\"00\"),'Turn Log'!$S$5:$S$5000,0)),\"\"))",
        "L": "=IF(A{r}=\"\",\"\",COUNTIFS('Turn Log'!$B$5:$B$5000,A{r},'Turn Log'!$X$5:$X$5000,1))",
        "M": "=IF(A{r}=\"\",\"\",COUNTIFS('Turn Log'!$B$5:$B$5000,A{r},'Turn Log'!$Y$5:$Y$5000,1))",
        "N": "=IF(A{r}=\"\",\"\",COUNTIFS('Turn Log'!$B$5:$B$5000,A{r},'Turn Log'!$W$5:$W$5000,1))",
        "O": "=IF(A{r}=\"\",\"\",IF(H{r}=\"\",\"No\",\"Yes\"))",
        "P": "=IF(A{r}=\"\",\"\",COUNTIFS('Stability Checks'!$A$5:$A$5000,A{r},'Stability Checks'!$C$5:$C$5000,\"Re-execution\",'Stability Checks'!$E$5:$E$5000,\"Pass\"))",
        "Q": "=IF(A{r}=\"\",\"\",COUNTIFS('Stability Checks'!$A$5:$A$5000,A{r},'Stability Checks'!$C$5:$C$5000,\"Re-prompt\",'Stability Checks'!$E$5:$E$5000,\"Pass\"))",
        "R": "=IF(A{r}=\"\",\"\",'Run Plan'!Q{r})",
    }
    for col_letter, template in formulas_by_col.items():
        ws[f"{col_letter}{row}"] = template.format(r=row)


def set_dashboard_formulas(ws) -> None:
    ws["B4"] = '=COUNTIF(\'Run Plan\'!$A$5:$A$5000,"?*")'
    ws["D4"] = '=COUNTIF(\'Run Summary\'!$O$5:$O$5000,"Yes")'
    ws["F4"] = '=IFERROR(AVERAGEIF(\'Run Summary\'!$H$5:$H$5000,">0"),"")'
    ws["B8"] = '=IFERROR(AVERAGEIF(\'Run Summary\'!$A$5:$A$5000,"?*",\'Run Summary\'!$M$5:$M$5000),"")'
    ws["D8"] = '=COUNTIF(\'Run Summary\'!$A$5:$A$5000,"?*")'
    ws["F8"] = '=COUNTIF(\'Stability Checks\'!$A$5:$A$5000,"?*")'


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Reset workbook to a manual-entry state: clear Run Plan, Turn Log, Stability Checks, and derived run rows while keeping reference sheets."
    )
    parser.add_argument("--workbook", required=True, help="Path to workbook/Experiment_ProgLang_Error_Propagation_repo.xlsx")
    parser.add_argument("--clear-tasks", action="store_true", help="Also clear Tasks rows if you want to choose tasks entirely by hand.")
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    wb = openpyxl.load_workbook(workbook_path)

    clear_rows(wb["Run Plan"], 5, RUN_PLAN_COLS)
    clear_rows(wb["Turn Log"], 5, TURN_LOG_COLS)
    clear_rows(wb["Stability Checks"], 5, STABILITY_COLS)

    if args.clear_tasks:
        clear_rows(wb["Tasks"], 5, TASK_COLS)

    run_summary = wb["Run Summary"]
    clear_rows(run_summary, 5, RUN_SUMMARY_COLS)
    set_run_summary_row_formulas(run_summary, 5)

    set_dashboard_formulas(wb["Dashboard"])

    mark_for_recalc(wb)
    wb.save(workbook_path)
    print("Workbook reset complete.")


if __name__ == "__main__":
    main()
