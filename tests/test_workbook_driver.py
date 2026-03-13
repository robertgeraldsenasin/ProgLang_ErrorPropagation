from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook

from errorprop_sql.workbook_driver import build_run_id, load_run_request, update_run_plan_row


REPO_ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_TEMPLATE = REPO_ROOT / "workbook" / "Experiment_ProgLang_Error_Propagation_repo.xlsx"


def test_build_run_id_uses_model_slug() -> None:
    assert build_run_id("local002", "GPT 5.3", "F2", 1) == "LOCAL002-GPT-5-3-F2-R1"


def test_load_run_request_defaults_and_updates(tmp_path: Path) -> None:
    workbook_path = tmp_path / "manual.xlsx"
    shutil.copy2(WORKBOOK_TEMPLATE, workbook_path)
    wb = load_workbook(workbook_path)
    ws = wb["Run Plan"]
    ws["C5"] = "local002"
    ws["E5"] = "gpt53-free"
    ws["G5"] = "F2"
    wb.save(workbook_path)

    req = load_run_request(workbook_path, row=5)
    assert req.run_id == "LOCAL002-GPT53-FREE-F2-R1"
    assert req.t_max == 4
    assert req.reasoning_mode == "manual-browser"

    update_run_plan_row(workbook_path, 5, {"run_id": req.run_id, "status": "Planned", "commit_hash": "abc123"})
    wb2 = load_workbook(workbook_path, data_only=False)
    ws2 = wb2["Run Plan"]
    assert ws2["A5"].value == req.run_id
    assert ws2["Q5"].value == "Planned"
    assert ws2["P5"].value == "abc123"
