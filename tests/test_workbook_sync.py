from __future__ import annotations

import json
import shutil
from pathlib import Path

from openpyxl import load_workbook

from errorprop_sql.workbook_sync import sync_output_to_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_TEMPLATE = REPO_ROOT / "workbook" / "Experiment_ProgLang_Error_Propagation_repo.xlsx"


def _write_jsonl(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row) + "\n")


def test_workbook_sync_appends_and_updates_rows(tmp_path: Path) -> None:
    workbook_path = tmp_path / "sync_test.xlsx"
    shutil.copy2(WORKBOOK_TEMPLATE, workbook_path)

    out_dir = tmp_path / "output"
    logs_dir = out_dir / "logs"

    run_id = "SMOKE-local001-gpt51-F1-r1"
    _write_jsonl(
        logs_dir / "run_plan.jsonl",
        [
            {
                "run_id": run_id,
                "batch_id": "SMOKE_A",
                "task_id": "local001",
                "db_name": "sqlite-demo",
                "model_id": "gpt-5.1",
                "model_snapshot": "gpt-5.1",
                "protocol_id": "F1",
                "temperature": 0,
                "reasoning_mode": "default",
                "T_max": 5,
                "replicate": 1,
                "operator": "tester",
                "planned_date": "2026-03-10",
                "record_file": "",
                "artifact_folder": "runs/SMOKE-local001-gpt51-F1-r1",
                "commit_hash": "abc123",
                "status": "In progress",
                "notes": "",
            },
            {
                "run_id": run_id,
                "batch_id": "SMOKE_A",
                "task_id": "local001",
                "db_name": "sqlite-demo",
                "model_id": "gpt-5.1",
                "model_snapshot": "gpt-5.1",
                "protocol_id": "F1",
                "temperature": 0,
                "reasoning_mode": "default",
                "T_max": 5,
                "replicate": 1,
                "operator": "tester",
                "planned_date": "2026-03-10",
                "record_file": "",
                "artifact_folder": "runs/SMOKE-local001-gpt51-F1-r1",
                "commit_hash": "abc123",
                "status": "Complete",
                "notes": "final_state=Pass; pass_turn=2",
            },
        ],
    )

    _write_jsonl(
        logs_dir / "turn_log.jsonl",
        [
            {
                "batch_id": "SMOKE_A",
                "run_id": run_id,
                "turn_no": 1,
                "task_id": "local001",
                "db_name": "sqlite-demo",
                "model_id": "gpt-5.1",
                "protocol_id": "F1",
                "prompt_file": "prompts/turn01.txt",
                "response_file": "responses/turn01.txt",
                "extracted_sql_file": "sql/turn01.sql",
                "prompt_hash": "hash1",
                "exec_ms": 14,
                "rows_pred": 1,
                "rows_gold": 1,
                "symdiff_rows": 2,
                "sqlite_error": "",
                "exec_state": "WrongResult",
                "feedback_file": "feedback/turn01.txt",
                "screenshot_file": "",
                "recording_timestamp": "10:01:00",
                "notes": "extraction=fenced_block",
                "review_flag": "",
            },
            {
                "batch_id": "SMOKE_A",
                "run_id": run_id,
                "turn_no": 2,
                "task_id": "local001",
                "db_name": "sqlite-demo",
                "model_id": "gpt-5.1",
                "protocol_id": "F1",
                "prompt_file": "prompts/turn02.txt",
                "response_file": "responses/turn02.txt",
                "extracted_sql_file": "sql/turn02.sql",
                "prompt_hash": "hash2",
                "exec_ms": 11,
                "rows_pred": 1,
                "rows_gold": 1,
                "symdiff_rows": 0,
                "sqlite_error": "",
                "exec_state": "Pass",
                "feedback_file": "feedback/turn02.txt",
                "screenshot_file": "",
                "recording_timestamp": "10:02:00",
                "notes": "extraction=fenced_block",
                "review_flag": "",
            },
        ],
    )

    _write_jsonl(
        logs_dir / "stability_checks.jsonl",
        [
            {
                "run_id": run_id,
                "pass_turn": 2,
                "check_type": "Re-execution",
                "repeat_no": 1,
                "outcome_state": "Pass",
                "stable_pass?": "Yes",
                "evidence_file": "screens/reexec1.png",
                "notes": "smoke",
            },
            {
                "run_id": run_id,
                "pass_turn": 2,
                "check_type": "Re-prompt",
                "repeat_no": 1,
                "outcome_state": "Pass",
                "stable_pass?": "Yes",
                "evidence_file": "responses/reprompt1.txt",
                "notes": "smoke",
            },
        ],
    )

    sync_output_to_workbook(workbook_path, out_dir)

    wb = load_workbook(workbook_path, data_only=False)

    run_plan_ws = wb["Run Plan"]
    turn_log_ws = wb["Turn Log"]
    stability_ws = wb["Stability Checks"]
    run_summary_ws = wb["Run Summary"]
    dashboard_ws = wb["Dashboard"]

    smoke_run_plan_rows = [r for r in range(5, run_plan_ws.max_row + 1) if run_plan_ws[f"A{r}"].value == run_id]
    assert len(smoke_run_plan_rows) == 1
    run_row = smoke_run_plan_rows[0]
    assert run_plan_ws[f"Q{run_row}"].value == "Complete"
    assert run_plan_ws[f"R{run_row}"].value == "final_state=Pass; pass_turn=2"

    smoke_turn_rows = [r for r in range(5, turn_log_ws.max_row + 1) if turn_log_ws[f"B{r}"].value == run_id]
    assert len(smoke_turn_rows) == 2
    assert turn_log_ws[f"Q{smoke_turn_rows[0]}"].value == "WrongResult"
    assert turn_log_ws[f"Q{smoke_turn_rows[1]}"].value == "Pass"
    assert "VLOOKUP" in str(turn_log_ws[f"R{smoke_turn_rows[0]}"].value)
    assert "$Q$5:$Q$5000" in str(turn_log_ws[f"U{smoke_turn_rows[1]}"].value)

    smoke_stability_rows = [r for r in range(5, stability_ws.max_row + 1) if stability_ws[f"A{r}"].value == run_id]
    assert len(smoke_stability_rows) == 2

    assert "MATCH(A" in str(run_summary_ws[f"J{run_row}"].value)
    assert run_summary_ws[f"A{run_row + 1}"].value is None
    assert 'COUNTIF(\'Run Summary\'!$A$5:$A$5000,"?*")' in str(dashboard_ws["D8"].value)
    assert 'COUNTIF(\'Run Plan\'!$A$5:$A$5000,"?*")' in str(dashboard_ws["B4"].value)


def test_workbook_sync_preserves_manual_run_plan_fields_when_log_row_is_blank(tmp_path: Path) -> None:
    workbook_path = tmp_path / "sync_blank_preserve.xlsx"
    shutil.copy2(WORKBOOK_TEMPLATE, workbook_path)

    wb = load_workbook(workbook_path)
    ws = wb["Run Plan"]
    ws["A5"] = "R-BLANK"
    ws["L5"] = "analyst"
    ws["N5"] = "recordings/run1.mp4"
    ws["P5"] = "abc123"
    ws["Q5"] = "Planned"
    wb.save(workbook_path)

    out_dir = tmp_path / "output"
    logs_dir = out_dir / "logs"
    _write_jsonl(
        logs_dir / "run_plan.jsonl",
        [
            {
                "run_id": "R-BLANK",
                "batch_id": "B1",
                "task_id": "local002",
                "db_name": "E_commerce",
                "model_id": "gpt53-free",
                "model_snapshot": "",
                "protocol_id": "F2",
                "temperature": 0,
                "reasoning_mode": "",
                "T_max": 4,
                "replicate": 1,
                "operator": "",
                "planned_date": "",
                "record_file": "",
                "artifact_folder": "runs/R-BLANK",
                "commit_hash": "",
                "status": "Complete",
                "notes": "",
            }
        ],
    )

    sync_output_to_workbook(workbook_path, out_dir)
    wb2 = load_workbook(workbook_path, data_only=False)
    ws2 = wb2["Run Plan"]
    assert ws2["L5"].value == "analyst"
    assert ws2["N5"].value == "recordings/run1.mp4"
    assert ws2["P5"].value == "abc123"
    assert ws2["Q5"].value == "Complete"
